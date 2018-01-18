from __future__ import absolute_import, division, print_function, unicode_literals
import arcpy

arcpy.env.overwriteOutput = True
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from matplotlib.backends.backend_pdf import PdfPages
import matplotlib.ticker as tick
import tempfile
from shutil import copyfile
import datetime
import os
import glob
import re
import xmltodict
from pylab import rcParams


rcParams['figure.figsize'] = 15, 10


def printmes(x):
    try:
        import arcpy
        arcpy.AddMessage(x)
        print(x)
    except ModuleNotFoundError:
        print(x)

def new_lev_imp(infile):
    with open(infile, "r") as fd:
        txt = fd.readlines()

    try:
        data_ind = txt.index('[Data]\n')
        # inst_info_ind = txt.index('[Instrument info from data header]\n')
        ch1_ind = txt.index('[CHANNEL 1 from data header]\n')
        ch2_ind = txt.index('[CHANNEL 2 from data header]\n')
        level = txt[ch1_ind + 1].split('=')[-1].strip()
        level_units = txt[ch1_ind + 2].split('=')[-1].strip()
        temp = txt[ch2_ind + 1].split('=')[-1].strip()
        temp_units = txt[ch2_ind + 2].split('=')[-1].strip()
        # serial_num = txt[inst_info_ind+1].split('=')[-1].strip().strip(".")
        # inst_num = txt[inst_info_ind+2].split('=')[-1].strip()
        # location = txt[inst_info_ind+3].split('=')[-1].strip()
        # start_time = txt[inst_info_ind+6].split('=')[-1].strip()
        # stop_time = txt[inst_info_ind+7].split('=')[-1].strip()

        df = pd.read_table(infile, parse_dates=[[0, 1]], sep='\s+', skiprows=data_ind + 2,
                           names=['Date', 'Time', level, temp],
                           skipfooter=1, engine='python')
        df.rename(columns={'Date_Time': 'DateTime'}, inplace=True)
        df.set_index('DateTime', inplace=True)

        printmes('Level units are {:}.'.format(level_units))
        if level_units == "feet" or level_units == "ft":
            df[level] = pd.to_numeric(df[level])
        elif level_units == "kpa":
            df[level] = pd.to_numeric(df[level]) * 0.33456
            printmes("Units in kpa, converting to ft...")
        elif level_units == "mbar":
            df[level] = pd.to_numeric(df[level]) * 0.0334552565551
        elif level_units == "psi":
            df[level] = pd.to_numeric(df[level]) * 2.306726
            printmes("Units in psi, converting to ft...")
        elif level_units == "m" or level_units == "meters":
            df[level] = pd.to_numeric(df[level]) * 3.28084
            printmes("Units in psi, converting to ft...")
        else:
            df[level] = pd.to_numeric(df[level])
            printmes("Unknown units, no conversion")

        if temp_units == 'Deg C' or temp_units == u'\N{DEGREE SIGN}' + u'C':
            df[temp] = df[temp]
        elif temp_units == 'Deg F' or temp_units == u'\N{DEGREE SIGN}' + u'F':
            printmes('Temp in F, converting to C')
            df[temp] = (df[temp] - 32.0) * 5.0 / 9.0

        return df
    except ValueError:
        printmes('File {:} has formatting issues'.format(infile))

def new_xle_imp(infile):
    """This function uses an exact file path to upload a xle transducer file.

    Args:
        infile (file):
            complete file path to input file

    Returns:
        A Pandas DataFrame containing the transducer data
    """
    # open text file
    with open(infile, "rb") as f:
        obj = xmltodict.parse(f, xml_attribs=True, encoding="ISO-8859-1")
    # navigate through xml to the data
    wellrawdata = obj['Body_xle']['Data']['Log']
    # convert xml data to pandas dataframe
    f = pd.DataFrame(wellrawdata)

    # CH 3 check
    try:
        ch3ID = obj['Body_xle']['Ch3_data_header']['Identification']
        f[str(ch3ID).title()] = f['ch3']
    except(KeyError, UnboundLocalError):
        pass

    # CH 2 manipulation
    ch2ID = obj['Body_xle']['Ch2_data_header']['Identification']
    f[str(ch2ID).title()] = f['ch2']
    ch2Unit = obj['Body_xle']['Ch2_data_header']['Unit']
    numCh2 = pd.to_numeric(f['ch2'])
    if ch2Unit == 'Deg C' or ch2Unit == u'\N{DEGREE SIGN}' + u'C':
        f[str(ch2ID).title()] = numCh2
    elif ch2Unit == 'Deg F' or ch2Unit == u'\N{DEGREE SIGN}' + u'F':
        printmes('Temp in F, converting to C')
        f[str(ch2ID).title()] = (numCh2 - 32) * 5 / 9

    # CH 1 manipulation
    ch1ID = obj['Body_xle']['Ch1_data_header']['Identification']  # Usually level
    ch1Unit = obj['Body_xle']['Ch1_data_header']['Unit']  # Usually ft
    unit = str(ch1Unit).lower()

    if unit == "feet" or unit == "ft":
        f[str(ch1ID).title()] = pd.to_numeric(f['ch1'])
    elif unit == "kpa":
        f[str(ch1ID).title()] = pd.to_numeric(f['ch1']) * 0.33456
        printmes("Units in kpa, converting to ft...")
    elif unit == "mbar":
        f[str(ch1ID).title()] = pd.to_numeric(f['ch1']) * 0.0334552565551
    elif unit == "psi":
        f[str(ch1ID).title()] = pd.to_numeric(f['ch1']) * 2.306726
        printmes("Units in psi, converting to ft...")
    elif unit == "m" or unit == "meters":
        f[str(ch1ID).title()] = pd.to_numeric(f['ch1']) * 3.28084
        printmes("Units in psi, converting to ft...")
    else:
        f[str(ch1ID).title()] = pd.to_numeric(f['ch1'])
        printmes("Unknown units, no conversion")

    # add extension-free file name to dataframe
    f['name'] = infile.split('\\').pop().split('/').pop().rsplit('.', 1)[0]
    # combine Date and Time fields into one field
    f['DateTime'] = pd.to_datetime(f.apply(lambda x: x['Date'] + ' ' + x['Time'], 1))
    f[str(ch1ID).title()] = pd.to_numeric(f[str(ch1ID).title()])
    f[str(ch2ID).title()] = pd.to_numeric(f[str(ch2ID).title()])

    try:
        ch3ID = obj['Body_xle']['Ch3_data_header']['Identification']
        f[str(ch3ID).title()] = pd.to_numeric(f[str(ch3ID).title()])
    except(KeyError, UnboundLocalError):
        pass

    f = f.reset_index()
    f = f.set_index('DateTime')
    f['Level'] = f[str(ch1ID).title()]
    f = f.drop(['Date', 'Time', '@id', 'ch1', 'ch2', 'index', 'ms'], axis=1)

    return f


def new_csv_imp(infile):
    """This function uses an exact file path to upload a csv transducer file.

    Args:
        infile (file):
            complete file path to input file

    Returns:
        A Pandas DataFrame containing the transducer data
    """
    f = pd.read_csv(infile, skiprows=1, parse_dates=[[0, 1]])
    # f = f.reset_index()
    f['DateTime'] = pd.to_datetime(f['Date_ Time'], errors='coerce')
    f = f[f.DateTime.notnull()]
    if ' Feet' in list(f.columns.values):
        f['Level'] = f[' Feet']
        f.drop([' Feet'], inplace=True, axis=1)
    elif 'Feet' in list(f.columns.values):
        f['Level'] = f['Feet']
        f.drop(['Feet'], inplace=True, axis=1)
    else:
        f['Level'] = f.iloc[:, 1]
    # Remove first and/or last measurements if the transducer was out of the water
    # f = dataendclean(f, 'Level')
    flist = f.columns.tolist()
    if ' Temp C' in flist:
        f['Temperature'] = f[' Temp C']
        f['Temp'] = f['Temperature']
        f.drop([' Temp C', 'Temperature'], inplace=True, axis=1)
    elif ' Temp F' in flist:
        f['Temperature'] = (f[' Temp F'] - 32) * 5 / 9
        f['Temp'] = f['Temperature']
        f.drop([' Temp F', 'Temperature'], inplace=True, axis=1)
    else:
        f['Temp'] = np.nan
    f.set_index(['DateTime'], inplace=True)
    f['date'] = f.index.to_julian_date().values
    f['datediff'] = f['date'].diff()
    f = f[f['datediff'] > 0]
    f = f[f['datediff'] < 1]
    # bse = int(pd.to_datetime(f.index).minute[0])
    # f = hourly_resample(f, bse)
    f.rename(columns={' Volts': 'Volts'}, inplace=True)
    f.drop([u'date', u'datediff', u'Date_ Time'], inplace=True, axis=1)
    return f


def dataendclean(df, x, inplace=False):
    """Trims off ends and beginnings of datasets that exceed 2.0 standard deviations of the first and last 30 values

    :param df: Pandas DataFrame
    :type df: pandas.core.frame.DataFrame
    :param x: Column name of data to be trimmed contained in df
    :type x: str
    :param inplace: if DataFrame should be duplicated
    :type inplace: bool

    :returns: df trimmed data
    :rtype: pandas.core.frame.DataFrame

    This function printmess a message if data are trimmed.
    """
    # Examine Mean Values
    if inplace:
        df = df
    else:
        df = df.copy()

    jump = df[abs(df.loc[:, x].diff()) > 1.0]
    try:
        for i in range(len(jump)):
            if jump.index[i] < df.index[50]:
                df = df[df.index > jump.index[i]]
                printmes("Dropped from beginning to " + str(jump.index[i]))
            if jump.index[i] > df.index[-50]:
                df = df[df.index < jump.index[i]]
                printmes("Dropped from end to " + str(jump.index[i]))
    except IndexError:
        printmes('No Jumps')
    return df


def new_trans_imp(infile):
    """This function uses an imports and cleans the ends of transducer file.

    Args:
        infile (file):
            complete file path to input file
        xle (bool):
            if true, then the file type should be xle; else it should be csv

    Returns:
        A Pandas DataFrame containing the transducer data
    """
    file_ext = os.path.splitext(infile)[1]
    try:
        if file_ext == '.xle':
            well = new_xle_imp(infile)
        elif file_ext == '.lev':
            well = new_lev_imp(infile)
        elif file_ext == '.csv':
            well = new_csv_imp(infile)
        else:
            printmes('filetype not recognized')
            pass
        return dataendclean(well,'Level')
    except AttributeError:
        printmes('Bad File')
        pass
    # Use `g[wellinfo[wellinfo['Well']==wellname]['closest_baro']]` to match the closest barometer to the data


def correct_be(site_number, well_table, welldata, be=None, meas='corrwl', baro='barometer'):
    if be:
        be = float(be)
    else:
        stdata = well_table[well_table['WellID'] == site_number]
        be = stdata['BaroEfficiency'].values[0]
    if be is None:
        be = 0
    else:
        be = float(be)

    if be == 0:
        welldata['BAROEFFICIENCYLEVEL'] = welldata[meas]
    else:
        welldata['BAROEFFICIENCYLEVEL'] = welldata[[meas, baro]].apply(lambda x: x[0] + be * x[1], 1)

    return welldata, be


def hourly_resample(df, bse=0, minutes=60):
    """
    resamples data to hourly on the hour
    Args:
        df:
            pandas dataframe containing time series needing resampling
        bse (int):
            base time to set; optional; default is zero (on the hour);
        minutes (int):
            sampling recurrence interval in minutes; optional; default is 60 (hourly samples)
    Returns:
        A Pandas DataFrame that has been resampled to every hour, at the minute defined by the base (bse)
    Description:
        see http://pandas.pydata.org/pandas-docs/dev/generated/pandas.DataFrame.resample.html for more info
        This function uses pandas powerful time-series manipulation to upsample to every minute, then downsample to every hour,
        on the hour.
        This function will need adjustment if you do not want it to return hourly samples, or iusgsGisf you are sampling more frequently than
        once per minute.
        see http://pandas.pydata.org/pandas-docs/stable/timeseries.html#offset-aliases
    """
    if int(str(pd.__version__).split('.')[0]) == 0 and int(
            str(pd.__version__).split('.')[1]) < 18:  # pandas versioning
        df = df.resample('1Min', how='first')
    else:
        # you can make this smaller to accomodate for a higher sampling frequency
        df = df.resample('1Min').first()

        # http://pandas.pydata.org/pandas-docs/dev/generated/pandas.Series.interpolate.html
    df = df.interpolate(method='time', limit=90)

    if int(str(pd.__version__).split('.')[0]) == 0 and int(
            str(pd.__version__).split('.')[1]) < 18:  # pandas versioning
        df = df.resample(str(minutes) + 'Min', closed='left', label='left', base=bse, how='first')
    else:
        # modify '60Min' to change the resulting frequency
        df = df.resample(str(minutes) + 'Min', closed='left', label='left', base=bse).first()
    return df


def well_baro_merge(wellfile, barofile, barocolumn='Level', wellcolumn='Level', outcolumn='corrwl',
                    vented=False,
                    sampint=60):
    """Remove barometric pressure from nonvented transducers.
    Args:
        wellfile (pd.DataFrame):
            Pandas DataFrame of water level data labeled 'Level'; index must be datetime
        barofile (pd.DataFrame):
            Pandas DataFrame barometric data labeled 'Level'; index must be datetime
        sampint (int):
            sampling interval in minutes; default 60

    Returns:
        wellbaro (Pandas DataFrame):
           corrected water levels with bp removed
    """

    # resample data to make sample interval consistent
    baro = hourly_resample(barofile, 0, sampint)
    well = hourly_resample(wellfile, 0, sampint)

    # reassign `Level` to reduce ambiguity
    baro = baro.rename(columns={barocolumn: 'barometer'})

    if 'TEMP' in baro.columns:
        baro.drop('TEMP', axis=1, inplace=True)

    # combine baro and well data for easy calculations, graphing, and manipulation
    wellbaro = pd.merge(well, baro, left_index=True, right_index=True, how='inner')

    wellbaro['dbp'] = wellbaro['barometer'].diff()
    wellbaro['dwl'] = wellbaro[wellcolumn].diff()
    first_well = wellbaro[wellcolumn][0]

    if vented:
        wellbaro[outcolumn] = wellbaro[wellcolumn]
    else:
        wellbaro[outcolumn] = wellbaro[['dbp', 'dwl']].apply(lambda x: x[1] - x[0], 1).cumsum() + first_well
    wellbaro.loc[wellbaro.index[0], outcolumn] = first_well
    return wellbaro


def imp_one_well(well_file, baro_file, man_startdate, man_start_level, man_endate, man_end_level,
                 conn_file_root,
                 wellid, be=None, well_table="UGGP.UGGPADMIN.UGS_NGWMN_Monitoring_Locations",
                 gw_reading_table="UGGP.UGGPADMIN.UGS_GW_reading", drift_tol=0.3, override=False):
    import arcpy
    arcpy.env.workspace = conn_file_root

    if os.path.splitext(well_file)[1] == '.xle':
        trans_type = 'Solinst'
    else:
        trans_type = 'Global Water'

    printmes('Trans type for well is {:}.'.format(trans_type))


    well = new_trans_imp(well_file)
    baro = new_trans_imp(baro_file)


    corrwl = well_baro_merge(well, baro, vented=(trans_type != 'Solinst'))

    if be:
        corrwl = correct_be(wellid, corrwl, be=be)
        corrwl['corrwl'] = corrwl['BAROEFFICIENCYLEVEL']

    stickup, well_elev = get_stickup_elev(wellid, well_table)

    man = pd.DataFrame(
        {'DateTime': [man_startdate, man_endate], 'MeasuredDTW': [man_start_level, man_end_level]}).set_index(
        'DateTime')
    printmes(man)
    man['Meas_GW_Elev'] = well_elev - (man['MeasuredDTW'] - stickup)

    man['MeasuredDTW'] = man['MeasuredDTW'] * -1

    dft = fix_drift(corrwl, man, meas='corrwl', manmeas='MeasuredDTW')
    drift = round(float(dft[1]['drift'].values[0]), 3)
    printmes('Drift for well {:} is {:}.'.format(wellid, drift))
    df = dft[0]

    rowlist, fieldnames = prepare_fieldnames(df, wellid, stickup, well_elev)

    if drift <= drift_tol:
        edit_table(rowlist, gw_reading_table, fieldnames)
        printmes('Well {:} successfully imported!'.format(wellid))
    elif override == 1:
        edit_table(rowlist, gw_reading_table, fieldnames)
        printmes('Override initiated. Well {:} successfully imported!'.format(wellid))
    else:
        printmes('Well {:} drift greater than tolerance!'.format(wellid))
    return df, man, be, drift


def prepare_fieldnames(df, wellid, stickup, well_elev, level='Level', dtw='DTW_WL'):
    """
    This function adds the necessary field names to import well data into the SDE database.
    :param df: pandas DataFrame of processed well data
    :param wellid: wellid (alternateID) of well in Stations Table
    :param level: raw transducer level from new_trans_imp, new_xle_imp, or new_csv_imp functions
    :param dtw: drift-corrected depth to water from fix_drift function
    :return: processed df with necessary field names for import
    """

    df['MEASUREDLEVEL'] = df[level]
    df['MEASUREDDTW'] = df[dtw] * -1
    df['DTWBELOWCASING'] = df['MEASUREDDTW']
    df['DTWBELOWGROUNDSURFACE'] = df['MEASUREDDTW'].apply(lambda x: x - stickup, 1)
    df['WATERELEVATION'] = df['DTWBELOWGROUNDSURFACE'].apply(lambda x: well_elev - x, 1)
    df['TAPE'] = 0
    df['LOCATIONID'] = wellid

    df.sort_index(inplace=True)

    fieldnames = ['READINGDATE', 'MEASUREDLEVEL', 'MEASUREDDTW', 'DRIFTCORRECTION',
                  'TEMP', 'LOCATIONID', 'DTWBELOWCASING', 'BAROEFFICIENCYLEVEL',
                  'DTWBELOWGROUNDSURFACE', 'WATERELEVATION', 'TAPE']

    if 'Temperature' in df.columns:
        df.rename(columns={'Temperature': 'TEMP'}, inplace=True)

    if 'TEMP' in df.columns:
        df['TEMP'] = df['TEMP'].apply(lambda x: np.round(x, 4), 1)
    else:
        df['TEMP'] = None

    if 'BAROEFFICIENCYLEVEL' in df.columns:
        pass
    else:
        df['BAROEFFICIENCYLEVEL'] = 0
    # subset bp df and add relevant fields
    df.index.name = 'READINGDATE'

    subset = df.reset_index()

    return subset, fieldnames


def find_extreme(site_number, gw_table="UGGP.UGGPADMIN.UGS_GW_reading", extma='max'):
    """
    Find extrema from a SDE table using query parameters
    :param table: SDE table to be queried
    :param site_number: LocationID of the site of interest
    :param extma: options are 'max' (default) or 'min'
    :return: read_max
    """
    import arcpy
    from arcpy import env
    env.overwriteOutput = True

    if extma == 'max':
        sort = 'DESC'
    else:
        sort = 'ASC'
    query = "LOCATIONID = '{:}'".format(site_number)
    field_names = ['READINGDATE', 'LOCATIONID', 'DTWBELOWGROUNDSURFACE', 'WATERELEVATION']
    sql_sn = ('TOP 1', 'ORDER BY READINGDATE {:}'.format(sort))
    # use a search cursor to iterate rows
    dateval, dtw, wlelev = [], [], []

    envtable = os.path.join(env.workspace, gw_table)

    with arcpy.da.SearchCursor(envtable, field_names, query, sql_clause=sql_sn) as search_cursor:
        # iterate the rows
        for row in search_cursor:
            dateval.append(row[0])
            dtw.append(row[1])
            wlelev.append(row[2])
    if len(dateval) < 1:
        return None, 0, 0
    else:
        return dateval[0], dtw[0], wlelev[0]

def getwellid(infile, wellinfo):
    """Specialized function that uses a well info table and file name to lookup a well's id number"""
    m = re.search("\d", getfilename(infile))
    s = re.search("\s", getfilename(infile))
    if m.start() > 3:
        wellname = getfilename(infile)[0:m.start()].strip().lower()
    else:
        wellname = getfilename(infile)[0:s.start()].strip().lower()
    wellid = wellinfo[wellinfo['Well'] == wellname]['WellID'].values[0]
    return wellname, wellid


def compilation(inputfile):
    """This function reads multiple xle transducer files in a directory and generates a compiled Pandas DataFrame.
    Args:
        inputfile (file):
            complete file path to input files; use * for wildcard in file name
    Returns:
        outfile (object):
            Pandas DataFrame of compiled data
    Example::
        >>> compilation('O:\\Snake Valley Water\\Transducer Data\\Raw_data_archive\\all\\LEV\\*baro*')
        picks any file containing 'baro'
    """

    # create empty dictionary to hold DataFrames
    f = {}

    # generate list of relevant files
    filelist = glob.glob(inputfile)

    # iterate through list of relevant files
    for infile in filelist:
        # run computations using lev files
        f[getfilename(infile)] = new_trans_imp(infile)
    # concatenate all of the DataFrames in dictionary f to one DataFrame: g
    g = pd.concat(f)
    # remove multiindex and replace with index=Datetime
    g = g.reset_index()
    g = g.set_index(['DateTime'])
    # drop old indexes
    g = g.drop(['level_0'], axis=1)
    # remove duplicates based on index then sort by index
    g['ind'] = g.index
    g.drop_duplicates(subset='ind', inplace=True)
    g.drop('ind', axis=1, inplace=True)
    g = g.sort_index()
    outfile = g
    return outfile


def getfilename(path):
    """This function extracts the file name without file path or extension

    Args:
        path (file):
            full path and file (including extension of file)

    Returns:
        name of file as string
    """
    return path.split('\\').pop().split('/').pop().rsplit('.', 1)[0]


def get_stickup_elev(site_number, well_table):
    wells = table_to_pandas_dataframe(well_table, field_names=['AltLocationID', 'Offset', 'Altitude'])
    stdata = wells[wells['AltLocationID'] == str(site_number)]
    stickup = float(stdata['Offset'].values[0])
    well_elev = float(stdata['Altitude'].values[0])
    return stickup, well_elev


def get_gw_elevs(site_number, well_table, manual, stable_elev=True):
    """
    Gets basic well parameters and most recent groundwater level data for a well id for dtw calculations.
    :param site_number: well site number in the site table
    :param manual: Pandas Dataframe of manual data
    :param table: pandas dataframe of site table;
    :param lev_table: groundwater level table; defaults to "UGGP.UGGPADMIN.UGS_GW_reading"
    :return: stickup, well_elev, be, maxdate, dtw, wl_elev
    """

    stdata = well_table[well_table['WellID'] == str(site_number)]
    man_sub = manual[manual['Location ID'] == int(site_number)]
    well_elev = float(stdata['Altitude'].values[0])

    if stable_elev:
        stickup = float(stdata['Offset'].values[0])
    else:
        stickup = man_sub['Current Stickup Height']

    # manual = manual['MeasuredDTW'].to_frame()
    man_sub.loc[:, 'MeasuredDTW'] = man_sub['Water Level (ft)'] * -1
    man_sub.loc[:, 'Meas_GW_Elev'] = man_sub['MeasuredDTW'].apply(lambda x: well_elev + (x + stickup), 1)

    return man_sub, stickup, well_elev


def get_field_names(table):
    read_descr = arcpy.Describe(table)
    field_names = []
    for field in read_descr.fields:
        field_names.append(field.name)
    field_names.remove('OBJECTID')
    return field_names


def table_to_pandas_dataframe(table, field_names=None, query=None, sql_sn=(None, None)):
    """
    Load data into a Pandas Data Frame for subsequent analysis.
    :param table: Table readable by ArcGIS.
    :param field_names: List of fields.
    :return: Pandas DataFrame object.
    """
    # if field names are not specified
    if not field_names:
        field_names = get_field_names(table)
    # create a pandas data frame
    df = pd.DataFrame(columns=field_names)

    # use a search cursor to iterate rows
    with arcpy.da.SearchCursor(table, field_names, query, sql_clause=sql_sn) as search_cursor:
        # iterate the rows
        for row in search_cursor:
            # combine the field names and row items together, and append them
            df = df.append(dict(zip(field_names, row)), ignore_index=True)

    # return the pandas data frame
    return df


def edit_table(df, gw_reading_table, fieldnames):
    """
    Edits SDE table by inserting new rows
    :param rowlist: pandas DataFrame converted to row list by df.values.tolist()
    :param gw_reading_table: sde table to edit
    :param fieldnames: field names that are being appended in order of appearance in dataframe or list row
    :return:
    """

    table_names = get_field_names(gw_reading_table)

    for name in fieldnames:
        if name not in table_names:
            fieldnames.remove(name)
            printmes("{:} not in {:} fieldnames!".format(name, gw_reading_table))

    if len(fieldnames) > 0:
        subset = df[fieldnames]
        rowlist = subset.values.tolist()

        arcpy.env.overwriteOutput = True
        edit = arcpy.da.Editor(arcpy.env.workspace)
        edit.startEditing(False, False)
        edit.startOperation()

        cursor = arcpy.da.InsertCursor(gw_reading_table, fieldnames)
        for j in range(len(rowlist)):
            cursor.insertRow(rowlist[j])

        del cursor
        edit.stopOperation()
        edit.stopEditing(True)
    else:
        printmes('No data imported!')


def simp_imp_well(well_table, file, baro_out, wellid, manual, stbl_elev=True,
                  gw_reading_table="UGGP.UGGPADMIN.UGS_GW_reading", drift_tol=0.3, override=False):
    # import well file
    well = new_trans_imp(file)

    file_ext = os.path.splitext(file)[1]
    if file_ext == '.xle':
        trans_type = 'Solinst'
    else:
        trans_type = 'Global Water'
    try:
        baroid = well_table.loc[wellid, 'BaroLoggerType']
        printmes('{:}'.format(baroid))
        corrwl = well_baro_merge(well, baro_out[str(baroid)], barocolumn='MEASUREDLEVEL',
                                      vented=(trans_type != 'Solinst'))
    except:
        corrwl = well_baro_merge(well, baro_out['9003'], barocolumn='MEASUREDLEVEL',
                                      vented=(trans_type != 'Solinst'))

    # be, intercept, r = clarks(corrwl, 'barometer', 'corrwl')
    # correct barometric efficiency
    wls, be = correct_be(wellid, well_table, corrwl)

    # get manual groundwater elevations
    # man, stickup, well_elev = self.get_gw_elevs(wellid, well_table, manual, stable_elev = stbl_elev)
    stdata = well_table[well_table['WellID'] == str(wellid)]
    man_sub = manual[manual['Location ID'] == int(wellid)]
    well_elev = float(stdata['Altitude'].values[0]) # Should be in feet

    if stbl_elev:
        if stdata['Offset'].values[0] is None:
            stickup = 0
            printmes('Well ID {:} missing stickup!'.format(wellid))
        else:
            stickup = float(stdata['Offset'].values[0])
    else:

        stickup = man_sub.loc[man_sub.last_valid_index(), 'Current Stickup Height']

    # manual = manual['MeasuredDTW'].to_frame()
    man_sub.loc[:, 'MeasuredDTW'] = man_sub['Water Level (ft)'] * -1
    man_sub.loc[:, 'Meas_GW_Elev'] = man_sub['MeasuredDTW'].apply(lambda x: float(well_elev) + (x + float(stickup)),
                                                                  1)
    printmes('Stickup: {:}, Well Elev: {:}'.format(stickup, well_elev))

    # fix transducer drift

    dft = fix_drift(wls, man_sub, meas='BAROEFFICIENCYLEVEL', manmeas='MeasuredDTW')
    drift = np.round(float(dft[1]['drift'].values[0]), 3)

    df = dft[0]
    df.sort_index(inplace=True)
    first_index = df.first_valid_index()

    # Get last reading at the specified location
    read_max, dtw, wlelev = find_extreme(wellid)

    printmes("Last database date is {:}. First transducer reading is on {:}.".format(read_max, first_index))

    rowlist, fieldnames = prepare_fieldnames(df, wellid, stickup, well_elev)

    if (read_max is None or read_max < first_index) and (drift < drift_tol):
        edit_table(rowlist, gw_reading_table, fieldnames)
        printmes(arcpy.GetMessages())
        printmes("Well {:} imported.".format(wellid))
    elif override and (drift < drift_tol):
        edit_table(rowlist, gw_reading_table, fieldnames)
        printmes(arcpy.GetMessages())
        printmes("Override Activated. Well {:} imported.".format(wellid))
    elif drift > drift_tol:
        printmes('Drift for well {:} exceeds tolerance!'.format(wellid))
    else:
        printmes('Dates later than import data for well {:} already exist!'.format(wellid))
        pass

    # except (ValueError, ZeroDivisionError):

    #   drift = -9999
    #    df = corrwl
    #    pass
    return rowlist, man_sub, be, drift


def fcl(df, dtObj):
    """Finds closest date index in a dataframe to a date object
    Args:
        df:
            DataFrame
        dtObj:
            date object

    taken from: http://stackoverflow.com/questions/15115547/find-closest-row-of-dataframe-to-given-time-in-pandas
    """
    return df.iloc[np.argmin(np.abs(pd.to_datetime(df.index) - dtObj))]  # remove to_pydatetime()


def fix_drift(well, manualfile, meas='Level', manmeas='MeasuredDTW', outcolname='DTW_WL'):
    """Remove transducer drift from nonvented transducer data. Faster and should produce same output as fix_drift_stepwise
    Args:
        well (pd.DataFrame):
            Pandas DataFrame of merged water level and barometric data; index must be datetime
        manualfile (pandas.core.frame.DataFrame):
            Pandas DataFrame of manual measurements
        meas (str):
            name of column in well DataFrame containing transducer data to be corrected
        manmeas (str):
            name of column in manualfile Dataframe containing manual measurement data
        outcolname (str):
            name of column resulting from correction
    Returns:
        wellbarofixed (pandas.core.frame.DataFrame):
            corrected water levels with bp removed
        driftinfo (pandas.core.frame.DataFrame):
            dataframe of correction parameters
    """
    # breakpoints = self.get_breakpoints(wellbaro, manualfile)
    breakpoints = []
    manualfile.index = pd.to_datetime(manualfile.index)
    manualfile.sort_index(inplace=True)

    for i in range(len(manualfile)):
        breakpoints.append(fcl(well, manualfile.index[i]).name)
    breakpoints = sorted(list(set(breakpoints)))

    bracketedwls, drift_features = {}, {}

    if well.index.name:
        dtnm = well.index.name
    else:
        dtnm = 'DateTime'
        well.index.name = 'DateTime'

    manualfile.loc[:, 'julian'] = manualfile.index.to_julian_date()
    for i in range(len(breakpoints) - 1):
        # Break up pandas dataframe time series into pieces based on timing of manual measurements
        bracketedwls[i] = well.loc[
            (well.index.to_datetime() > breakpoints[i]) & (well.index.to_datetime() < breakpoints[i + 1])]
        df = bracketedwls[i]
        if len(df) > 0:
            df.loc[:, 'julian'] = df.index.to_julian_date()

            last_trans = df.loc[df.index[-1], meas]  # last transducer measurement
            first_trans = df.loc[df.index[0], meas]  # first transducer measurement

            last_man = fcl(manualfile, breakpoints[i + 1])  # first manual measurment
            first_man = fcl(manualfile, breakpoints[i])  # last manual mesurement

            # intercept of line = value of first manual measurement
            b = first_trans - first_man[manmeas]

            # slope of line = change in difference between manual and transducer over time;
            drift = ((last_trans - last_man[manmeas]) - b)
            m = drift / (last_man['julian'] - first_man['julian'])

            # datechange = amount of time between manual measurements
            df.loc[:, 'datechange'] = df['julian'].apply(lambda x: x - df.loc[df.index[0], 'julian'], 1)

            # bracketedwls[i].loc[:, 'wldiff'] = bracketedwls[i].loc[:, meas] - first_trans
            # apply linear drift to transducer data to fix drift; flipped x to match drift
            df.loc[:, 'DRIFTCORRECTION'] = df['datechange'].apply(lambda x: m * x, 1)
            df[outcolname] = df[meas] - (df['DRIFTCORRECTION'] + b)

            drift_features[i] = {'begining': first_man, 'end': last_man, 'intercept': b, 'slope': m,
                                 'first_meas': first_man[manmeas], 'last_meas': last_man[manmeas],
                                 'drift': drift}
        else:
            pass

    wellbarofixed = pd.concat(bracketedwls)
    wellbarofixed.reset_index(inplace=True)
    wellbarofixed.set_index(dtnm, inplace=True)
    drift_info = pd.DataFrame(drift_features).T

    return wellbarofixed, drift_info


def xle_head_table(folder):
    """Creates a Pandas DataFrame containing header information from all xle files in a folder
    Args:
        folder (directory):
            folder containing xle files
    Returns:
        A Pandas DataFrame containing the transducer header data
    Example::
        >>> xle_head_table('C:/folder_with_xles/')
    """
    # open text file
    df = {}
    for infile in os.listdir(folder):

        # get the extension of the input file
        filename, filetype = os.path.splitext(folder + infile)
        basename = os.path.basename(folder + infile)
        if filetype == '.xle':
            # open text file
            with open(folder + infile, "rb") as f:
                d = xmltodict.parse(f, xml_attribs=True, encoding="ISO-8859-1")
            # navigate through xml to the data
            data = list(d['Body_xle']['Instrument_info_data_header'].values()) + list(
                d['Body_xle']['Instrument_info'].values())
            cols = list(d['Body_xle']['Instrument_info_data_header'].keys()) + list(
                d['Body_xle']['Instrument_info'].keys())

            df[basename[:-4]] = pd.DataFrame(data=data, index=cols).T
    allwells = pd.concat(df)
    allwells.index = allwells.index.droplevel(1)
    allwells.index.name = 'filename'
    allwells['trans type'] = 'Solinst'
    allwells['fileroot'] = allwells.index
    allwells['full_filepath'] = allwells['fileroot'].apply(lambda x: folder + x + '.xle', 1)

    return allwells


def csv_info_table(folder):
    csv = {}
    files = [f for f in os.listdir(folder) if os.path.isfile(os.path.join(folder, f))]
    field_names = ['filename', 'Start_time', 'Stop_time']
    df = pd.DataFrame(columns=field_names)
    for file in files:
        fileparts = os.path.basename(file).split('.')
        filetype = fileparts[1]
        basename = fileparts[0]
        if filetype == 'csv':
            try:
                cfile = {}
                csv[basename] = new_csv_imp(os.path.join(folder, file))
                cfile['Battery_level'] = int(round(csv[basename].loc[csv[basename]. \
                                                   index[-1], 'Volts'] / csv[basename]. \
                                                   loc[csv[basename].index[0], 'Volts'] * 100, 0))
                cfile['Sample_rate'] = (csv[basename].index[1] - csv[basename].index[0]).seconds * 100
                cfile['filename'] = basename
                cfile['fileroot'] = basename
                cfile['full_filepath'] = os.path.join(folder, file)
                cfile['Start_time'] = csv[basename].first_valid_index()
                cfile['Stop_time'] = csv[basename].last_valid_index()
                cfile['Location'] = ' '.join(basename.split(' ')[:-1])
                cfile['trans type'] = 'Global Water'
                df = df.append(cfile, ignore_index=True)
            except:
                pass
    df.set_index('filename', inplace=True)
    return df, csv


def upload_bp_data(df, site_number, return_df=False, gw_reading_table="UGGP.UGGPADMIN.UGS_GW_reading"):
    import arcpy

    df.sort_index(inplace=True)
    first_index = df.first_valid_index()

    # Get last reading at the specified location
    read_max, dtw, wlelev = find_extreme(site_number)

    if read_max is None or read_max < first_index:

        df['MEASUREDLEVEL'] = df['Level']
        df['TAPE'] = 0
        df['LOCATIONID'] = site_number

        df.sort_index(inplace=True)

        fieldnames = ['READINGDATE', 'MEASUREDLEVEL', 'TEMP', 'LOCATIONID', 'TAPE']

        if 'Temperature' in df.columns:
            df.rename(columns={'Temperature': 'TEMP'}, inplace=True)

        if 'TEMP' in df.columns:
            df['TEMP'] = df['TEMP'].apply(lambda x: np.round(x, 4), 1)
        else:
            df['TEMP'] = None

        df.index.name = 'READINGDATE'

        subset = df.reset_index()

        edit_table(subset, gw_reading_table, fieldnames)

        if return_df:
            return df

    else:
        printmes('Dates later than import data for this station already exist!')
        pass


def get_location_data(site_number, enviro, first_date=None, last_date=None, limit=None,
                      gw_reading_table="UGGP.UGGPADMIN.UGS_GW_reading"):
    arcpy.env.workspace = enviro
    if not first_date:
        first_date = datetime.datetime(1900, 1, 1)
    elif type(first_date) == str:
        try:
            datetime.datetime.strptime(first_date, '%m/%d/%Y')
        except:
            first_date = datetime.datetime(1900, 1, 1)
    # Get last reading at the specified location
    if not last_date or last_date > datetime.datetime.now():
        last_date = datetime.datetime.now()

    query_txt = "LOCATIONID = '{:}' and (READINGDATE >= '{:%m/%d/%Y}' and READINGDATE <= '{:%m/%d/%Y}')"
    query = query_txt.format(site_number, first_date, last_date + datetime.timedelta(days=1))
    printmes(query)
    sql_sn = (limit, 'ORDER BY READINGDATE ASC')

    fieldnames = get_field_names(gw_reading_table)

    readings = table_to_pandas_dataframe(gw_reading_table, fieldnames, query, sql_sn)
    readings.set_index('READINGDATE', inplace=True)
    if len(readings) == 0:
        printmes('No Records for location {:}'.format(site_number))
    return readings

class baroimport(object):
    def __init__(self):
        self.sde_conn = None
        self.wellid = None
        self.xledir = None
        self.well_files = None
        self.wellname = None
        self.welldict = None
        self.filedict = None
        self.man_file = None
        self.save_location = None
        self.should_plot = None
        self.chart_out = None
        self.tol = None
        self.stbl = None
        self.ovrd = None
        self.toexcel = None
        self.baro_comp_file = None
        self.to_import = None
        self.idget = None

    def many_baros(self):
        """Used by the MultBarometerImport tool to import multiple wells into the SDE"""
        arcpy.env.workspace = self.sde_conn


        self.xledir = self.xledir+r"\\"

        # upload barometric pressure data
        df = {}


        if self.should_plot:
            pdf_pages = PdfPages(self.chart_out)


        for b in range(len(self.wellid)):

            sitename = self.filedict[self.well_files[b]]
            altid = self.idget[sitename]
            printmes([b,altid,sitename])
            df[altid] = new_trans_imp(self.xledir + self.well_files[b])
            printmes("Importing {:} ({:})".format(sitename, altid))

            if self.to_import:
                upload_bp_data(df[altid], altid)
                printmes('Barometer {:} ({:}) Imported'.format(sitename, altid))

            if self.toexcel:
                from openpyxl import load_workbook
                if b == 0:
                    writer = pd.ExcelWriter(self.xledir + '/wells.xlsx')
                    df[altid].to_excel(writer, sheet_name='{:}_{:}'.format(sitename, b))
                    writer.save()
                    writer.close()
                else:
                    book = load_workbook(self.xledir + '/wells.xlsx')
                    writer = pd.ExcelWriter(self.xledir + '/wells.xlsx', engine='openpyxl')
                    writer.book = book
                    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
                    df[altid].to_excel(writer, sheet_name='{:}_{:}'.format(sitename, b))
                    writer.save()
                    writer.close()

            if self.should_plot:
                # plot data
                df[altid].set_index('READINGDATE', inplace=True)
                y1 = df[altid]['WATERELEVATION'].values
                y2 = df[altid]['barometer'].values
                x1 = df[altid].index.values
                x2 = df[altid].index.values

                fig, ax1 = plt.subplots()

                ax1.plot(x1, y1, color='blue', label='Water Level Elevation')
                ax1.set_ylabel('Water Level Elevation', color='blue')
                ax1.set_ylim(min(df[altid]['WATERELEVATION']), max(df[altid]['WATERELEVATION']))
                y_formatter = tick.ScalarFormatter(useOffset=False)
                ax1.yaxis.set_major_formatter(y_formatter)
                ax2 = ax1.twinx()
                ax2.set_ylabel('Barometric Pressure (ft)', color='red')
                ax2.plot(x2, y2, color='red', label='Barometric pressure (ft)')
                h1, l1 = ax1.get_legend_handles_labels()
                h2, l2 = ax2.get_legend_handles_labels()
                ax1.legend(h1 + h2, l1 + l2, loc=3)
                plt.xlim(df[altid].first_valid_index() - datetime.timedelta(days=3),
                         df[altid].last_valid_index() + datetime.timedelta(days=3))
                plt.title('Well: {:}'.format(sitename))
                pdf_pages.savefig(fig)
                plt.close()

            h = pd.read_csv(self.baro_comp_file, index_col=0, header=0, parse_dates=True)
            g = pd.concat([h, df[altid]])
            # remove duplicates based on index then sort by index
            g['ind'] = g.index
            g.drop_duplicates(subset='ind', inplace=True)
            g.drop('ind', axis=1, inplace=True)
            g = g.sort_index()
            os.remove(self.baro_comp_file)
            g.to_csv(self.baro_comp_file)


        if self.should_plot:
            pdf_pages.close()

        return


class wellimport(object):
    def __init__(self):
        self.sde_conn = None
        self.well_file = None
        self.baro_file = None
        self.man_startdate = None
        self.man_enddate = None
        self.man_start_level = None
        self.man_end_level = None
        self.wellid = None
        self.xledir = None
        self.well_files = None
        self.wellname = None
        self.welldict = None
        self.filedict = None
        self.man_file = None
        self.save_location = None
        self.should_plot = None
        self.chart_out = None
        self.tol = None
        self.stbl = None
        self.ovrd = None
        self.toexcel = None
        self.baro_comp_file = None
        self.to_import = None
        self.idget = None

    def read_xle(self):
        well = new_xle_imp(self.well_file)
        well.to_csv(self.save_location)
        return


    def one_well(self):
        """Used in SingleTransducerImport Class.  This tool leverages the imp_one_well function to load a single well
        into the UGS SDE"""
        arcpy.env.workspace = self.sde_conn
        loc_table = "UGGP.UGGPADMIN.UGS_NGWMN_Monitoring_Locations"

        loc_names = [str(row[0]) for row in arcpy.da.SearchCursor(loc_table, 'LocationName')]
        loc_ids = [str(row[0]) for row in arcpy.da.SearchCursor(loc_table, 'AltLocationID')]

        iddict = dict(zip(loc_names, loc_ids))

        if self.man_startdate in ["#", "", None]:
            self.man_startdate, self.man_start_level, wlelev = find_extreme(self.wellid)

        df, man, be, drift = imp_one_well(self.well_file, self.baro_file, self.man_startdate,
                                               self.man_start_level, self.man_enddate,
                                               self.man_end_level, self.sde_conn, iddict.get(self.wellid),
                                               drift_tol=self.tol, override=self.ovrd)

        if self.should_plot:
            # plot data
            pdf_pages = PdfPages(self.chart_out)
            y1 = df['WATERELEVATION'].values
            y2 = df['barometer'].values
            x1 = df.index.values
            x2 = df.index.values

            x4 = man.index
            y4 = man['Meas_GW_Elev']
            fig, ax1 = plt.subplots()
            ax1.scatter(x4, y4, color='purple')
            ax1.plot(x1, y1, color='blue', label='Water Level Elevation')
            ax1.set_ylabel('Water Level Elevation', color='blue')
            ax1.set_ylim(min(df['WATERELEVATION']), max(df['WATERELEVATION']))
            y_formatter = tick.ScalarFormatter(useOffset=False)
            ax1.yaxis.set_major_formatter(y_formatter)
            ax2 = ax1.twinx()
            ax2.set_ylabel('Barometric Pressure (ft)', color='red')
            ax2.plot(x2, y2, color='red', label='Barometric pressure (ft)')
            h1, l1 = ax1.get_legend_handles_labels()
            h2, l2 = ax2.get_legend_handles_labels()
            ax1.legend(h1 + h2, l1 + l2, loc=3)
            plt.xlim(df.first_valid_index() - datetime.timedelta(days=3),
                     df.last_valid_index() + datetime.timedelta(days=3))
            plt.title('Well: {:}  Drift: {:}  Baro. Eff.: {:}'.format(self.wellid, drift, be))
            pdf_pages.savefig(fig)
            plt.close()
            pdf_pages.close()

        printmes('Well Imported!')
        printmes(arcpy.GetMessages())
        return


    def remove_bp(self):

        well = self.new_trans_imp(self.well_file)
        baro = self.new_trans_imp(self.baro_file)

        df = self.well_baro_merge(well, baro, barocolumn='Level', wellcolumn='Level', outcolumn='corrwl', vented=False,
                                  sampint=60)

        df.to_csv(self.save_location)

    def remove_bp_drift(self):

        well = self.new_trans_imp(self.well_file)
        baro = self.new_trans_imp(self.baro_file)

        man = pd.DataFrame(
            {'DateTime': [self.man_startdate, self.man_enddate],
             'MeasuredDTW': [self.man_start_level * -1, self.man_end_level * -1]}).set_index('DateTime')

        corrwl = self.well_baro_merge(well, baro, barocolumn='Level', wellcolumn='Level', outcolumn='corrwl',
                                      vented=False,
                                      sampint=60)

        dft = self.fix_drift(corrwl, man, meas='corrwl', manmeas='MeasuredDTW')
        drift = round(float(dft[1]['drift'].values[0]), 3)

        printmes("Drift is {:} feet".format(drift))
        dft[0].to_csv(self.save_location)

        if self.should_plot:
            pdf_pages = PdfPages(self.chart_out)

            # plot data
            df = dft[0]
            y1 = df['DTW_WL'].values
            y2 = df['barometer'].values
            x1 = df.index.values
            x2 = df.index.values

            x4 = man.index
            y4 = man['MeasuredDTW']
            fig, ax1 = plt.subplots()
            plt.xticks(rotation=70)
            ax1.scatter(x4, y4, color='purple')
            ax1.plot(x1, y1, color='blue', label='Water Level')
            ax1.set_ylabel('Depth to Water (ft)', color='blue')
            ax1.set_ylim(min(y1), max(y1))
            y_formatter = tick.ScalarFormatter(useOffset=False)
            ax1.yaxis.set_major_formatter(y_formatter)
            ax2 = ax1.twinx()
            ax2.set_ylabel('Barometric Pressure (ft)', color='red')
            ax2.plot(x2, y2, color='red', label='Barometric pressure (ft)')
            h1, l1 = ax1.get_legend_handles_labels()
            h2, l2 = ax2.get_legend_handles_labels()
            ax1.legend(h1 + h2, l1 + l2, loc=3)
            plt.xlim(df.first_valid_index() - datetime.timedelta(days=3),
                     df.last_valid_index() + datetime.timedelta(days=3))

            pdf_pages.savefig(fig)
            plt.close()
            pdf_pages.close()

    def get_ftype(self, x):
        if x[1] == 'Solinst':
            ft = '.xle'
        else:
            ft = '.csv'
        return self.filedict.get(x[0] + ft)

    def many_wells(self):
        """Used by the MultTransducerImport tool to import multiple wells into the SDE"""
        arcpy.env.workspace = self.sde_conn
        loc_table = "UGGP.UGGPADMIN.UGS_NGWMN_Monitoring_Locations"

        # create empty dataframe to house well data
        field_names = ['LocationID', 'LocationName', 'LocationType', 'LocationDesc', 'AltLocationID', 'Altitude',
                       'AltitudeUnits', 'WellDepth', 'SiteID', 'Offset', 'LoggerType', 'BaroEfficiency',
                       'BaroEfficiencyStart', 'BaroLoggerType']
        df = pd.DataFrame(columns=field_names)
        # populate dataframe with data from SDE well table
        search_cursor = arcpy.da.SearchCursor(loc_table, field_names)
        for row in search_cursor:
            # combine the field names and row items together, and append them
            df = df.append(dict(zip(field_names, row)), ignore_index=True)
        df.dropna(subset=['AltLocationID'], inplace=True)

        # create temp directory and populate it with relevant files
        file_extension = []
        dirpath = tempfile.mkdtemp(suffix=r'\\')
        for file in self.well_files:
            copyfile(os.path.join(self.xledir, file), os.path.join(dirpath, file))
            file_extension.append(os.path.splitext(file)[1])

        # examine and tabulate header information from files

        if '.xle' in file_extension and '.csv' in file_extension:
            xles = xle_head_table(dirpath)
            printmes('xles examined')
            csvs = csv_info_table(dirpath)
            printmes('csvs examined')
            file_info_table = pd.concat([xles, csvs[0]])
        elif '.xle' in file_extension:
            xles = xle_head_table(dirpath)
            printmes('xles examined')
            file_info_table = xles
        elif '.csv' in file_extension:
            csvs = csv_info_table(dirpath)
            printmes('csvs examined')
            file_info_table = csvs[0]

        # combine header table with the sde table
        file_info_table['WellName'] = file_info_table[['fileroot', 'trans type']].apply(lambda x: self.get_ftype(x), 1)
        well_table = pd.merge(file_info_table, df, right_on='LocationName', left_on='WellName', how='left')
        well_table.set_index('AltLocationID', inplace=True)
        well_table['WellID'] = well_table.index
        well_table.dropna(subset=['WellName'], inplace=True)
        well_table.to_csv(self.xledir + '/file_info_table.csv')
        printmes("Header Table with well information created at {:}/file_info_table.csv".format(self.xledir))
        maxtime = max(pd.to_datetime(well_table['Stop_time']))
        mintime = min(pd.to_datetime(well_table['Start_time']))
        printmes("Data span from {:} to {:}.".format(mintime, maxtime))

        # upload barometric pressure data
        baro_out = {}
        baros = well_table[well_table['LocationType'] == 'Barometer']

        #lastdate = maxtime + datetime.timedelta(days=1)
        lastdate = None

        if len(baros) < 1:
            baros = [9024,9025,9027,9049,9061,9003]
            for baro in baros:
                try:
                    baro_out[str(baro)] = get_location_data(baro, self.sde_conn, first_date=mintime,
                                                            last_date=lastdate)
                    printmes('Barometer {:} data download success'.format(baro))
                except:
                    printmes('Barometer {:} Data not available'.format(baro))
                    pass

        else:
            for b in range(len(baros)):
                barline = baros.iloc[b, :]
                df = new_trans_imp(barline['full_filepath'])
                upload_bp_data(df, baros.index[b])
                baro_out[baros.index[b]] = get_location_data(baros.index[b], self.sde_conn, first_date=mintime,
                                                             last_date= lastdate)
                printmes('Barometer {:} ({:}) Imported'.format(barline['LocationName'], baros.index[b]))

        # upload manual data from csv file
        manl = pd.read_csv(self.man_file, index_col="DateTime")

        if self.should_plot:
            pdf_pages = PdfPages(self.chart_out)

        # import well data
        wells = well_table[well_table['LocationType'] == 'Well']
        for i in range(len(wells)):
            well_line = wells.iloc[i, :]
            printmes("Importing {:} ({:})".format(well_line['LocationName'],wells.index[i]))

            df, man, be, drift = simp_imp_well(well_table, well_line['full_filepath'], baro_out, wells.index[i],
                                                    manl, stbl_elev=self.stbl, drift_tol=float(self.tol), override=self.ovrd)
            printmes(arcpy.GetMessages())
            printmes('Drift for well {:} is {:}.'.format(well_line['LocationName'], drift))
            printmes("Well {:} complete.\n---------------".format(well_line['LocationName']))

            if self.toexcel:
                from openpyxl import load_workbook
                if i == 0:
                    writer = pd.ExcelWriter(self.xledir + '/wells.xlsx')
                    df.to_excel(writer, sheet_name='{:}_{:%Y%m}'.format(well_line['LocationName'], maxtime))
                    writer.save()
                    writer.close()
                else:
                    book = load_workbook(self.xledir + '/wells.xlsx')
                    writer = pd.ExcelWriter(self.xledir + '/wells.xlsx', engine='openpyxl')
                    writer.book = book
                    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
                    df.to_excel(writer, sheet_name='{:}_{:%Y%m}'.format(well_line['LocationName'], maxtime))
                    writer.save()
                    writer.close()

            if self.should_plot:
                # plot data
                df.set_index('READINGDATE', inplace=True)
                y1 = df['WATERELEVATION'].values
                y2 = df['barometer'].values
                x1 = df.index.values
                x2 = df.index.values

                x4 = man.index
                y4 = man['Meas_GW_Elev']
                fig, ax1 = plt.subplots()
                ax1.scatter(x4, y4, color='purple')
                ax1.plot(x1, y1, color='blue', label='Water Level Elevation')
                ax1.set_ylabel('Water Level Elevation', color='blue')
                ax1.set_ylim(min(df['WATERELEVATION']), max(df['WATERELEVATION']))
                y_formatter = tick.ScalarFormatter(useOffset=False)
                ax1.yaxis.set_major_formatter(y_formatter)
                ax2 = ax1.twinx()
                ax2.set_ylabel('Barometric Pressure (ft)', color='red')
                ax2.plot(x2, y2, color='red', label='Barometric pressure (ft)')
                h1, l1 = ax1.get_legend_handles_labels()
                h2, l2 = ax2.get_legend_handles_labels()
                ax1.legend(h1 + h2, l1 + l2, loc=3)
                plt.xlim(df.first_valid_index() - datetime.timedelta(days=3),
                         df.last_valid_index() + datetime.timedelta(days=3))
                plt.title('Well: {:}  Drift: {:}  Baro. Eff.: {:}'.format(well_line['LocationName'], drift, be))
                pdf_pages.savefig(fig)
                plt.close()

        if self.should_plot:
            pdf_pages.close()

        return